import glob
import os
import tkinter
from tkinter import messagebox
from tkinter import ttk

import openpyxl
import openpyxl as ox
import pandas as pd
import pandasql as ps
from openpyxl.styles import Protection, PatternFill
from openpyxl.worksheet.protection import SheetProtection
from pandasql import sqldf

filepath = os.getcwd() + "/DataBase.xlsx"


def color_rows(ExcelFilePath):
    wb = openpyxl.load_workbook(ExcelFilePath)

    # Select the worksheet you want to work with
    ws = wb['Sheet1']

    # Define the cell range to check for the cell values
    cell_range = ws['A2':'K1000']

    # Define the fill color to apply to the rows that meet the condition
    fill_colors = {
        'Attendance': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Light red
        'Behavior': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # Light yellow
        'Evaluation': PatternFill(start_color='B6D7A8', end_color='B6D7A8', fill_type='solid'),  # Light green
        'Final': PatternFill(start_color='C7CEEA', end_color='C7CEEA', fill_type='solid'),  # Light purple
        'Tasks': PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid'),  # Light pink
        'Q1': PatternFill(start_color='B4A7D6', end_color='B4A7D6', fill_type='solid'),  # Light violet
        'Q2': PatternFill(start_color='A2C4C9', end_color='A2C4C9', fill_type='solid'),  # Light blue
        'Q3': PatternFill(start_color='F7B6D2', end_color='F7B6D2', fill_type='solid')  # Light magenta
    }
    # Loop through the rows in the cell range and check the cell values
    for row in cell_range:
        if row[6].value == 'Attendance':
            fill = fill_colors['Attendance']
            for cell in row:
                cell.fill = fill
        elif row[6].value == 'Behavior':
            fill = fill_colors['Behavior']
            for cell in row:
                cell.fill = fill
        elif row[6].value == 'Evaluation':
            fill = fill_colors['Evaluation']
            for cell in row:
                cell.fill = fill
        elif row[6].value == 'Final':
            fill = fill_colors['Final']
            for cell in row:
                cell.fill = fill
        elif row[6].value == 'Tasks':
            fill = fill_colors['Tasks']
            for cell in row:
                cell.fill = fill
        elif row[6].value == 'Q1':
            fill = fill_colors['Q1']
            for cell in row:
                cell.fill = fill
        elif row[6].value == 'Q2':
            fill = fill_colors['Q2']
            for cell in row:
                cell.fill = fill
        elif row[6].value == 'Q3':
            fill = fill_colors['Q3']
            for cell in row:
                cell.fill = fill
    worksheet = wb.active
    max_col = worksheet.max_column
    max_row = worksheet.max_row

    # Loop through all cells in the sheet and apply borders to them
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                            bottom=openpyxl.styles.Side(style='thin'))
            cell.border = border
    # Save the changes to the Excel file
    # worksheet.protection.sheet = False
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(11)].protection = Protection(locked=False)
    # worksheet.protection.sheet = True
    wb.save(ExcelFilePath)

def combine_subjects_sheets():
    folder_path = 'Subjects_Sheets/' + TermFolder_combobox.get()
    if os.path.exists(folder_path):
        combined_data = pd.DataFrame()

        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    df = pd.read_excel(file_path)
                    combined_data = combined_data.append(df, ignore_index=True)
                    print(file_path)

        output_path = f'Combined_Data_{TermFolder_combobox.get()}.xlsx'
        combined_data.to_excel(output_path, index=False)
        tkinter.messagebox.showinfo(title='Sheets Combined successfully',
                                    message='Sheet Created with at the following path:' + os.getcwd() + f'/Combined_Data_{TermFolder_combobox.get()}.xlsx')

    else:
        tkinter.messagebox.showwarning(title="Error",
                                       message="Folder doesn't exist, please make sure thet the selected folder exists under " + os.getcwd() + "/Subjects_Sheets and contains the subjects sheet for Language Types")


def Create_sheets_per_subject():
    term = TermExtract_combobox.get()
    df = pd.read_excel(output_entry.get() + ".xlsx")
    distinct_values = sqldf("SELECT DISTINCT SubjectId FROM df")
    progress_bar.start()
    i = 0.0
    for value in distinct_values['SubjectId']:
        subjectName = sqldf(f"SELECT DISTINCT SubjectNameEn FROM df WHERE SubjectId = '{value}'")
        subjectName2 = subjectName.to_string(index=1).replace("  SubjectNameEn\n0    ", "").replace(" ", "")

        GradeId = sqldf(f"SELECT DISTINCT GradeId FROM df WHERE SubjectId = '{value}'")
        GradeId2 = GradeId.to_string(index=1).replace("GradeId\n0", "").replace(" ", "")
        # print(GradeId2)

        LanguageType = sqldf(f"SELECT DISTINCT LanguageType FROM df WHERE SubjectId = '{value}'")
        LanguageType2 = LanguageType.to_string(index=1).replace("LanguageType\n0", "").replace(" ", "")
        # print(LanguageType2)
        # print(subjectName2)

        selected_rows = sqldf(f"SELECT * FROM df WHERE SubjectId = '{value}' ORDER BY QuizType ASC")

        if not os.path.exists("Subjects_Sheets"):
            os.makedirs("Subjects_Sheets")
        if not os.path.exists(f"Subjects_Sheets/Term_{term}"):
            os.makedirs(f"Subjects_Sheets/Term_{term}")
        if not os.path.exists(f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}"):
            os.makedirs(f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}")
        if not os.path.exists(f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}"):
            os.makedirs(f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}")

        selected_rows.to_excel(
            f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}/{subjectName2}_Id_{value}.xlsx",
            index=False)
        workbook = openpyxl.load_workbook(
            f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}/{subjectName2}_Id_{value}.xlsx")
        worksheet = workbook.active
        # for col in range(1, 11):
        # worksheet.column_dimensions[openpyxl.utils.get_column_letter(11)].protection = Protection(locked=False)
        # worksheet.protection.sheet = True
        worksheet.auto_filter.ref = "A1:J1000"
        worksheet.protection = SheetProtection(autoFilter=True)
        # worksheet.protection.set_password('FGS')

        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3

        workbook.save(
            f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}/{subjectName2}_Id_{value}.xlsx")
        color_rows(f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}/{subjectName2}_Id_{value}.xlsx")

        print(
            "Created:" + f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}/{subjectName2}_Id_{value}.xlsx")
        # color_rows(f"Subjects_Sheets/Term_{term}/Grade_{GradeId2}/Language Type_{LanguageType2}/{subjectName2}_Id_{value}.xlsx")
        # print(subjectName)
        progress_bar['value'] = i
        progress_bar.update()
        i = i + 0.41
    progress_bar.stop()

    tkinter.messagebox.showinfo(title='Sheets Created successfully',
                                message='Sheet Created with at the following folder:' + os.getcwd() + '/Subjects_Sheets')


def fill_exams_sheet():
    accepted = accept_var2.get()
    if accepted == "Accepted":
        DataBase = glob.glob(filepath)
        Term = TermExtract_combobox.get()
        outputfile = output_entry.get()

        # loop over the list of csv files

        for f in DataBase:
            # read the csv file
            grades = pd.read_excel(f, sheet_name="Grades")
            students = pd.read_excel(f, sheet_name="Students")
            subjects = pd.read_excel(f, sheet_name="Subjects")
            exams = pd.read_excel(f, sheet_name="Exams")
            teachers = pd.read_excel(f, sheet_name="Teachers")

            query = '''
            SELECT *, '{Term}' AS Term, '1' AS Month, '' AS Mark
            FROM (
                select s."Id" "StudentId","StudentNameEn", s."GradeId",sb."Id" "SubjectId","SubjectNameEn"
                      , g."QuizType",s."LanguageType"
                from students s
                join subjects sb on s."GradeId" = sb."GradeId" 
                                and s."LanguageType" = sb."LanguageType"
                                and s."Class" = sb."Class"
                join grades g on g."Id" = sb."GradeId" 
            ) AS original_query
            UNION ALL
            SELECT *, '{Term}' AS Term, '2' AS Month, '' AS Mark
            FROM (
                select s."Id" "StudentId","StudentNameEn", s."GradeId",sb."Id" "SubjectId","SubjectNameEn"
                      , g."QuizType",s."LanguageType"
                from students s
                join subjects sb on s."GradeId" = sb."GradeId" 
                                and s."LanguageType" = sb."LanguageType"
                                and s."Class" = sb."Class"
                join grades g on g."Id" = sb."GradeId" 
            ) AS repeated_query_2
            UNION ALL
            SELECT *, '{Term}' AS Term, '3' AS Month, '' AS Mark
            FROM (
                select s."Id" "StudentId","StudentNameEn", s."GradeId",sb."Id" "SubjectId","SubjectNameEn"
                      , g."QuizType",s."LanguageType"
                from students s
                join subjects sb on s."GradeId" = sb."GradeId" 
                                and s."LanguageType" = sb."LanguageType"
                                and s."Class" = sb."Class"
                join grades g on g."Id" = sb."GradeId" 
            ) AS repeated_query_3
            ORDER BY QuizType DESC, SubjectId DESC
            '''.format(Term=Term)

            query_output = ps.sqldf(query)
            query_output = ps.sqldf("SELECT * FROM query_output WHERE Month NOT IN (1, 2) OR QuizType != 'Final'")
            query_output = ps.sqldf("SELECT * FROM query_output WHERE NOT (Month = 3 AND QuizType IN ('Q1', 'Q2', 'Q3', 'Attendance', 'Participation', 'Evaluation', 'Behavior', 'Tasks'))")
            print(query_output)
            query_output.to_excel(outputfile + '.xlsx')
        tkinter.messagebox.showinfo(title='Regenerate Exams Sheet',
                                    message='Sheet Created with at the following path:' + os.getcwd() + '/' + outputfile + '.xlsx')
        start_button["state"] = 'normal'
    else:
        tkinter.messagebox.showwarning(title="Error", message="You need to confirm your inputs")


def isfloat(num):
    if num.isalpha():
        return False
    else:
        return True


def formatNumber(num):
    return num


def get_student_data():
    get_subject_Id_button["state"] = 'disabled'
    DB_StudentName.set('')
    DB_GradeId.set('')
    DB_LanguageType.set('')
    DB_ClassId.set('')
    SubjectName_combobox['values'] = ''

    StudentId = StudentId_entry.get()
    if StudentId:
        # filepath = os.getcwd()+"/DataBase_Old.xlsx"
        # filepath = "DataBase.xlsx"
        df = pd.read_excel(filepath, sheet_name="Students")
        StudentName = df.query('Id == ' + StudentId)['StudentNameEn']
        LanguageType = df.query('Id == ' + StudentId)['LanguageType']
        GradeId = df.query('Id == ' + StudentId)['GradeId']
        Class = df.query('Id == ' + StudentId)['Class']
        if StudentName.values:
            DB_StudentName.set(StudentName.values[0])
            DB_GradeId.set(GradeId.values[0])
            DB_LanguageType.set(LanguageType.values[0])
            DB_ClassId.set(Class.values[0])
            dff = pd.read_excel(filepath, sheet_name="Subjects")
            print(
                'GradeId == ' + DB_GradeId.get() + ' and LanguageType == ' + DB_LanguageType.get() + ' and Class == ' + DB_ClassId.get())
            subjects = dff.query(
                'GradeId == ' + DB_GradeId.get() + ' and LanguageType == \'' + DB_LanguageType.get() + '\' and Class == \'' + DB_ClassId.get() + '\'')[
                'SubjectNameEn']
            subjectsList = subjects.values.tolist()
            SubjectName_combobox['values'] = subjectsList
            get_subject_Id_button["state"] = 'normal'
        else:
            tkinter.messagebox.showerror(title="Error", message="No student found, Please enter correct ID")
    else:
        tkinter.messagebox.showerror(title="Error", message="Please enter a student ID")


def get_subject_Id():
    DB_SubjectId.set('')
    SubjectName = SubjectName_combobox.get()
    GradeID = DB_GradeId.get()
    LanguageType = DB_LanguageType.get()
    Class = DB_ClassId.get()
    # filepath = os.getcwd()+"/DataBase_Old.xlsx"
    # filepath = "DataBase.xlsx"
    df = pd.read_excel(filepath, sheet_name="Subjects")
    subjectId = df.query(
        'SubjectNameEn == \'' + SubjectName + '\' and GradeId == ' + GradeID + ' and LanguageType == \'' + LanguageType + '\' and Class == \'' + Class + '\'')[
        'Id']
    print(
        'SubjectNameEn == \'' + SubjectName + '\' and GradeId == ' + GradeID + ' and LanguageType == \'' + LanguageType + '\' and Class == \'' + Class + '\'')
    # print(subjectId)
    if subjectId.values:
        DB_SubjectId.set(subjectId.values[0])
        if DB_SubjectId.get().isnumeric():
            get_Current_Student_Marks_button['state'] = 'normal'
        else:
            tkinter.messagebox.showerror(title="Error",
                                         message="There was an issue with the DB, please open the DB and refresh it then try again")
    else:
        tkinter.messagebox.showerror(title="Error",
                                     message="Please select a subject, then click get subject ID")


def get_Current_Student_Marks():
    DB_Attendance.set("")
    DB_Behavior.set("")
    DB_Evaluation.set("")
    DB_Participation.set("")
    DB_Q1.set("")
    DB_Q2.set("")
    DB_Q3.set("")
    # DB_Tasks.set("")
    # DB_Final.set("")
    StudentId = StudentId_entry.get()
    SubjectId = DB_SubjectId.get()
    Term = Term_combobox.get()
    Month = Month_Combobox.get()
    if not Term or not Month:
        tkinter.messagebox.showerror(title="Error", message="Please select month and term")
    else:
        enter_data_button['state'] = 'normal'
        # filepath = os.getcwd()+"/DataBase_Old.xlsx"
        # filepath = "DataBase.xlsx"
        df = pd.read_excel(filepath, sheet_name="Exams")
        Quizes_for_Subject = df.query(
            'StudentId == ' + StudentId + ' and SubjectId == ' + SubjectId + ' and Term == ' + Term + ' and Month == ' + Month)[
            'QuizType']
        StudentMarks = df.query(
            'StudentId == ' + StudentId + ' and SubjectId == ' + SubjectId + ' and Term == ' + Term + ' and Month == ' + Month)
        for QuizType in Quizes_for_Subject:
            if QuizType == 'Attendance':
                AttendanceMark = StudentMarks.query('QuizType == \'Attendance\'')['Mark']
                DB_Attendance.set(formatNumber(AttendanceMark.values[0]))
                if AttendanceMark.size > 1:
                    tkinter.messagebox.showwarning(title="Error",
                                                   message="The selected student has Duplicate values for Attendance, Please review the DB")
                    enter_data_button['state'] = 'disabled'
            if QuizType == 'Behavior':
                BehaviorMark = StudentMarks.query('QuizType == \'Behavior\'')['Mark']
                DB_Behavior.set(formatNumber(BehaviorMark.values[0]))
                if BehaviorMark.size > 1:
                    tkinter.messagebox.showwarning(title="Error",
                                                   message="The selected student has Duplicate values for Behavior, Please review the DB")
                    enter_data_button['state'] = 'disabled'
            if QuizType == 'Evaluation':
                EvaluationMark = StudentMarks.query('QuizType == \'Evaluation\'')['Mark']
                DB_Evaluation.set(formatNumber(EvaluationMark.values[0]))
                if EvaluationMark.size > 1:
                    tkinter.messagebox.showwarning(title="Error",
                                                   message="The selected student has Duplicate values for Evaluation, Please review the DB")
                    enter_data_button['state'] = 'disabled'
            if QuizType == 'Participation':
                ParticipationMark = StudentMarks.query('QuizType == \'Participation\'')['Mark']
                DB_Participation.set(formatNumber(ParticipationMark.values[0]))
                if ParticipationMark.size > 1:
                    tkinter.messagebox.showwarning(title="Error",
                                                   message="The selected student has Duplicate values for Participation, Please review the DB")
                    enter_data_button['state'] = 'disabled'
            if QuizType == 'Q1':
                Q1Mark = StudentMarks.query('QuizType == \'Q1\'')['Mark']
                DB_Q1.set(formatNumber(Q1Mark.values[0]))
                if Q1Mark.size > 1:
                    tkinter.messagebox.showwarning(title="Error",
                                                   message="The selected student has Duplicate values for Q1, Please review the DB")
                    enter_data_button['state'] = 'disabled'
            if QuizType == 'Q2':
                Q2Mark = StudentMarks.query('QuizType == \'Q2\'')['Mark']
                DB_Q2.set(formatNumber(Q2Mark.values[0]))
                if Q2Mark.size > 1:
                    tkinter.messagebox.showwarning(title="Error",
                                                   message="The selected student has Duplicate values for Q2, Please review the DB")
                    enter_data_button['state'] = 'disabled'
            if QuizType == 'Q3':
                Q3Mark = StudentMarks.query('QuizType == \'Q3\'')['Mark']
                DB_Q3.set(formatNumber(Q3Mark.values[0]))
                if Q3Mark.size > 1:
                    tkinter.messagebox.showwarning(title="Error",
                                                   message="The selected student has Duplicate values for Q3 or row doesn't exist, Please review the DB")
                    enter_data_button['state'] = 'disabled'
            # if QuizType == 'Tasks':
            #     TasksMark = StudentMarks.query('QuizType == \'Tasks\'')['Mark']
            #     DB_Tasks.set(formatNumber(TasksMark.values[0]))
            # if QuizType == 'Final':
            #     FinalMark = StudentMarks.query('QuizType == \'Final\'')['Mark']
            #     DB_Final.set(formatNumber(FinalMark.values[0]))
        if not DB_Q3.get() or not DB_Q2.get() or not DB_Q1.get() or not DB_Participation.get() or not DB_Behavior.get() or not DB_Attendance.get() or not DB_Evaluation.get():
            tkinter.messagebox.showwarning(title="Error",
                                           message="The selected student has no rows for some quizes, please review the DB")
            enter_data_button['state'] = 'disabled'


def enter_data():
    accepted = accept_var.get()
    if isfloat(DB_Attendance.get()) and isfloat(DB_Behavior.get()) and isfloat(DB_Evaluation.get()) and isfloat(
            DB_Participation.get()) and isfloat(DB_Q1.get()) and isfloat(DB_Q2.get()) and isfloat(DB_Q3.get()):
        if accepted == "Accepted":
            # filepath = os.getcwd()+"/DataBase_Old.xlsx"
            # filepath = "DataBase.xlsx"

            StudentId = StudentId_entry.get()
            SubjectId = DB_SubjectId.get()
            Term = Term_combobox.get()
            Month = Month_Combobox.get()
            df = pd.read_excel(filepath, sheet_name="Exams")
            Quizes_for_Subject = df.query(
                'StudentId == ' + StudentId + ' and SubjectId == ' + SubjectId + ' and Term == ' + Term + ' and Month == ' + Month)[
                'QuizType']
            StudentMarks = df.query(
                'StudentId == ' + StudentId + ' and SubjectId == ' + SubjectId + ' and Term == ' + Term + ' and Month == ' + Month)
            # print(Quizes_for_Subject)
            # QuizesNumber = Quizes_for_Subject.size
            wb = ox.load_workbook(filepath)
            ws = wb['Exams']
            col_no = df.columns.get_loc("Mark") + 1
            for QuizType in Quizes_for_Subject:
                if QuizType == 'Attendance':
                    AttendanceMark = StudentMarks.query('QuizType == \'Attendance\'')['Mark']
                    rowNumber = AttendanceMark.index.values[0] + 2
                    ws.cell(row=rowNumber, column=col_no).value = DB_Attendance.get()
                if QuizType == 'Behavior':
                    BehaviorMark = StudentMarks.query('QuizType == \'Behavior\'')['Mark']
                    rowNumber = BehaviorMark.index.values[0] + 2
                    ws.cell(row=rowNumber, column=col_no).value = DB_Behavior.get()
                if QuizType == 'Evaluation':
                    EvaluationMark = StudentMarks.query('QuizType == \'Evaluation\'')['Mark']
                    rowNumber = EvaluationMark.index.values[0] + 2
                    ws.cell(row=rowNumber, column=col_no).value = DB_Evaluation.get()
                if QuizType == 'Participation':
                    ParticipationMark = StudentMarks.query('QuizType == \'Participation\'')['Mark']
                    rowNumber = ParticipationMark.index.values[0] + 2
                    ws.cell(row=rowNumber, column=col_no).value = DB_Participation.get()
                if QuizType == 'Q1':
                    Q1Mark = StudentMarks.query('QuizType == \'Q1\'')['Mark']
                    rowNumber = Q1Mark.index.values[0] + 2
                    ws.cell(row=rowNumber, column=col_no).value = DB_Q1.get()
                if QuizType == 'Q2':
                    Q2Mark = StudentMarks.query('QuizType == \'Q2\'')['Mark']
                    rowNumber = Q2Mark.index.values[0] + 2
                    ws.cell(row=rowNumber, column=col_no).value = DB_Q2.get()
                if QuizType == 'Q3':
                    Q3Mark = StudentMarks.query('QuizType == \'Q3\'')['Mark']
                    rowNumber = Q3Mark.index.values[0] + 2
                    ws.cell(row=rowNumber, column=col_no).value = DB_Q3.get()
            wb.save(filepath)
            tkinter.messagebox.showinfo(title='Marks Entry',
                                        message='Your inputs are loaded into the database successfully')
            StudentId_entry.delete(0, tkinter.END)
            StudentName_value.delete(0, tkinter.END)
            Class_value.delete(0, tkinter.END)
            GradeId_value.delete(0, tkinter.END)
            LanguageType_value.delete(0, tkinter.END)
            SubjectName_combobox.delete(0, tkinter.END)
            Month_Combobox.delete(0, tkinter.END)
            SubjectId_entry.delete(0, tkinter.END)
            Term_combobox.delete(0, tkinter.END)
            Q1_entry.delete(0, tkinter.END)
            Q2_entry.delete(0, tkinter.END)
            Q3_entry.delete(0, tkinter.END)
            Attendance_entry.delete(0, tkinter.END)
            Behavior_entry.delete(0, tkinter.END)
            Evaluation_entry.delete(0, tkinter.END)
            Participation_entry.delete(0, tkinter.END)
            get_subject_Id_button["state"] = 'disabled'
            get_Current_Student_Marks_button['state'] = 'disabled'
            enter_data_button['state'] = 'disabled'


        else:
            tkinter.messagebox.showwarning(title="Error", message="You need to confirm your inputs")
    else:
        tkinter.messagebox.showwarning(title="Error",
                                       message="Please enter only numbers in the marks fields, if there is any value contains \"NaN\", remove it and set it with empty or your input")


def clear_all():
    get_subject_Id_button["state"] = 'disabled'
    get_Current_Student_Marks_button["state"] = 'disabled'
    StudentId_entry.delete(0, "end")
    DB_StudentName.set('')
    DB_GradeId.set('')
    DB_LanguageType.set('')
    DB_ClassId.set('')
    SubjectName_combobox['values'] = ''
    SubjectName_combobox.delete(0, "end")
    DB_SubjectId.set('')
    DB_Attendance.set("")
    DB_Behavior.set("")
    DB_Evaluation.set("")
    DB_Participation.set("")
    DB_Q1.set("")
    DB_Q2.set("")
    DB_Q3.set("")
    # DB_Tasks.set("")
    # DB_Final.set("")

    StudentId_entry.delete(0, tkinter.END)
    StudentName_value.delete(0, tkinter.END)
    Class_value.delete(0, tkinter.END)
    GradeId_value.delete(0, tkinter.END)
    LanguageType_value.delete(0, tkinter.END)
    SubjectName_combobox.delete(0, tkinter.END)
    Month_Combobox.delete(0, tkinter.END)
    SubjectId_entry.delete(0, tkinter.END)
    Term_combobox.delete(0, tkinter.END)
    Q1_entry.delete(0, tkinter.END)
    Q2_entry.delete(0, tkinter.END)
    Q3_entry.delete(0, tkinter.END)
    Attendance_entry.delete(0, tkinter.END)
    Behavior_entry.delete(0, tkinter.END)
    Evaluation_entry.delete(0, tkinter.END)
    Participation_entry.delete(0, tkinter.END)
    get_subject_Id_button["state"] = 'disabled'
    get_Current_Student_Marks_button['state'] = 'disabled'
    enter_data_button['state'] = 'disabled'


def get_Exams_per_student():
    # get_subject_Id_button2["state"] = 'disabled'
    DB_StudentName2.set('')
    DB_GradeId2.set('')
    DB_LanguageType2.set('')
    DB_ClassId2.set('')
    SubjectName_listbox.delete(0, 'end')
    SubjectId_listbox.delete(0, 'end')
    Quizes_listbox.delete(0, 'end')
    FullMarks_listbox.delete(0, 'end')
    RowsCount1_1.set('')
    RowsCount1_2.set('')
    RowsCount1_3.set('')
    RowsCount2_1.set('')
    RowsCount2_2.set('')
    RowsCount2_3.set('')
    # SubjectName_listbox['values'] = ''

    StudentId = StudentId_entry2.get()
    if StudentId:
        # filepath = os.getcwd()+"/DataBase_Old.xlsx"
        # filepath = "DataBase.xlsx"
        df = pd.read_excel(filepath, sheet_name="Students")
        StudentName = df.query('Id == ' + StudentId)['StudentNameEn']
        LanguageType = df.query('Id == ' + StudentId)['LanguageType']
        GradeId = df.query('Id == ' + StudentId)['GradeId']
        Class = df.query('Id == ' + StudentId)['Class']
        if StudentName.values:
            DB_StudentName2.set(StudentName.values[0])
            DB_GradeId2.set(GradeId.values[0])
            DB_LanguageType2.set(LanguageType.values[0])
            DB_ClassId2.set(Class.values[0])
            dff = pd.read_excel(filepath, sheet_name="Subjects")
            print(
                'GradeId == ' + DB_GradeId2.get() + ' and LanguageType == ' + DB_LanguageType2.get() + ' and Class == ' + DB_ClassId2.get())
            subjects = dff.query(
                'GradeId == ' + DB_GradeId2.get() + ' and LanguageType == \'' + DB_LanguageType2.get() + '\' and Class == \'' + DB_ClassId2.get() + '\'')[
                'SubjectNameEn']
            subjectsIds = dff.query(
                'GradeId == ' + DB_GradeId2.get() + ' and LanguageType == \'' + DB_LanguageType2.get() + '\' and Class == \'' + DB_ClassId2.get() + '\'')[
                'Id']
            subjectsList = subjects.values.tolist()
            subjectsIDsList = subjectsIds.values.tolist()
            SubjectName_listbox.insert(0, *subjectsList)
            SubjectId_listbox.insert(0, *subjectsIDsList)

            dff2 = pd.read_excel(filepath, sheet_name="Grades")
            QuizesTypes = dff2.query('Id == ' + DB_GradeId2.get())['QuizType']
            FullMarks = dff2.query('Id == ' + DB_GradeId2.get())['Full Mark']
            QuizesTypesList = QuizesTypes.values.tolist()
            FullMarksList = FullMarks.values.tolist()
            Quizes_listbox.insert(0, *QuizesTypesList)
            FullMarks_listbox.insert(0, *FullMarksList)
            RowsCount.set(len(subjectsList) * len(QuizesTypesList))

            dff3 = pd.read_excel(filepath, sheet_name="Exams")
            Rows1_1 = dff3.query('StudentId == ' + StudentId + ' and Term == 1 and Month == 1')['Id']
            Rows1_2 = dff3.query('StudentId == ' + StudentId + ' and Term == 1 and Month == 2')['Id']
            Rows1_3 = dff3.query('StudentId == ' + StudentId + ' and Term == 1 and Month == 3')['Id']
            Rows2_1 = dff3.query('StudentId == ' + StudentId + ' and Term == 2 and Month == 1')['Id']
            Rows2_2 = dff3.query('StudentId == ' + StudentId + ' and Term == 2 and Month == 2')['Id']
            Rows2_3 = dff3.query('StudentId == ' + StudentId + ' and Term == 2 and Month == 3')['Id']
            RowsCount1_1.set(Rows1_1.size)
            RowsCount1_2.set(Rows1_2.size)
            RowsCount1_3.set(Rows1_3.size)
            RowsCount2_1.set(Rows2_1.size)
            RowsCount2_2.set(Rows2_2.size)
            RowsCount2_3.set(Rows2_3.size)

        else:
            tkinter.messagebox.showerror(title="Error", message="No student found, Please enter correct ID")
    else:
        tkinter.messagebox.showerror(title="Error", message="Please enter a student ID")


# User Interface--------------------------------------------------------------------------------------------------------
window = tkinter.Tk()
tabControl = ttk.Notebook(window)
tab1 = ttk.Frame(tabControl)

tabControl.add(tab1, text='Student Marks')
tabControl.pack(expand=1, fill="both")

window.title("Student Marks Entry Form")

frame = tkinter.Frame(window)
frame.pack()

# Students Tab----------------------------------------------------------------------------------------------------------
user_info_frame = tkinter.LabelFrame(tab1, text="Student Data")
user_info_frame.grid(row=0, column=0, padx=20, pady=10)

StudentId_label = tkinter.Label(user_info_frame, text="Student ID")
StudentId_label.grid(row=0, column=0)
StudentName_label = tkinter.Label(user_info_frame, text="Student_Name")
StudentName_label.grid(row=0, column=1)

# Button
St_Data_button = tkinter.Button(user_info_frame, text="Get Student Data", command=get_student_data)
St_Data_button.grid(row=2, column=0, sticky="news", padx=20, pady=10)

StudentId_entry = tkinter.Entry(user_info_frame)
DB_StudentName = tkinter.StringVar()
DB_StudentName.set("")
StudentName_value = tkinter.Entry(user_info_frame, textvariable=DB_StudentName, state="disabled")
StudentId_entry.grid(row=1, column=0)
StudentName_value.grid(row=1, column=1)

GradeId_label = tkinter.Label(user_info_frame, text="Grade")
DB_GradeId = tkinter.StringVar()
DB_GradeId.set("")
GradeId_value = tkinter.Entry(user_info_frame, textvariable=DB_GradeId, state="disabled")
GradeId_label.grid(row=0, column=2)
GradeId_value.grid(row=1, column=2)

Class_label = tkinter.Label(user_info_frame, text="Class")
DB_ClassId = tkinter.StringVar()
DB_ClassId.set("")
Class_value = tkinter.Entry(user_info_frame, textvariable=DB_ClassId, state="disabled")
Class_label.grid(row=2, column=1)
Class_value.grid(row=3, column=1)

LanguageType_label = tkinter.Label(user_info_frame, text="LanguageType")
DB_LanguageType = tkinter.StringVar()
DB_LanguageType.set("")
LanguageType_value = tkinter.Entry(user_info_frame, textvariable=DB_LanguageType, state="disabled")
LanguageType_label.grid(row=2, column=2)
LanguageType_value.grid(row=3, column=2)

Month_label = tkinter.Label(user_info_frame, text="Month")
Month_Combobox = ttk.Combobox(user_info_frame, values=["1", "2", "3"], state="readonly")
Month_label.grid(row=2, column=3)
Month_Combobox.grid(row=3, column=3)

SubjectName_label = tkinter.Label(user_info_frame, text="SubjectName")
SubjectName_combobox = ttk.Combobox(user_info_frame, values=[], state="readonly")
SubjectName_label.grid(row=0, column=3)
SubjectName_combobox.grid(row=1, column=3)

# Button
get_subject_Id_button = tkinter.Button(user_info_frame, text="Get Subject ID", command=get_subject_Id, state='disabled')
get_subject_Id_button.grid(row=0, column=4, sticky="news", padx=20, pady=10)
# SubjectId_label = tkinter.Label(user_info_frame, text="Subject ID")

DB_SubjectId = tkinter.StringVar()
DB_SubjectId.set("")
SubjectId_entry = tkinter.Entry(user_info_frame, textvariable=DB_SubjectId, state="disabled")
# SubjectId_label.grid(row=0, column=4)
SubjectId_entry.grid(row=1, column=4)

Term_label = tkinter.Label(user_info_frame, text="Term")
Term_combobox = ttk.Combobox(user_info_frame, values=["1", "2"], state="readonly")
Term_label.grid(row=2, column=4)
Term_combobox.grid(row=3, column=4)

for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# Saving Marks Inputs
marks_frame = tkinter.LabelFrame(tab1, text="Student Marks and Notes")
marks_frame.grid(row=1, column=0, sticky="news", padx=20, pady=10)

# Button
get_Current_Student_Marks_button = tkinter.Button(marks_frame, text="Get Current Student Marks",
                                                  command=get_Current_Student_Marks, state='disabled')
get_Current_Student_Marks_button.grid(row=0, column=0, sticky="news", padx=20, pady=10)

Q1_label = tkinter.Label(marks_frame, text="Q1")
DB_Q1 = tkinter.StringVar()
DB_Q1.set("")
Q1_entry = tkinter.Entry(marks_frame, textvariable=DB_Q1)
Q1_label.grid(row=1, column=0)
Q1_entry.grid(row=2, column=0)

Q2_label = tkinter.Label(marks_frame, text="Q2")
DB_Q2 = tkinter.StringVar()
DB_Q2.set("")
Q2_entry = tkinter.Entry(marks_frame, textvariable=DB_Q2)
Q2_label.grid(row=1, column=1)
Q2_entry.grid(row=2, column=1)

Q3_label = tkinter.Label(marks_frame, text="Q3")
DB_Q3 = tkinter.StringVar()
DB_Q3.set("")
Q3_entry = tkinter.Entry(marks_frame, textvariable=DB_Q3)
Q3_label.grid(row=1, column=2)
Q3_entry.grid(row=2, column=2)

Attendance_label = tkinter.Label(marks_frame, text="Attendance")
DB_Attendance = tkinter.StringVar()
DB_Attendance.set("")
Attendance_entry = tkinter.Entry(marks_frame, textvariable=DB_Attendance)
Attendance_label.grid(row=1, column=3)
Attendance_entry.grid(row=2, column=3)

Behavior_label = tkinter.Label(marks_frame, text="Behavior")
DB_Behavior = tkinter.StringVar()
DB_Behavior.set("")
Behavior_entry = tkinter.Entry(marks_frame, textvariable=DB_Behavior)
Behavior_label.grid(row=3, column=0)
Behavior_entry.grid(row=4, column=0)

Evaluation_label = tkinter.Label(marks_frame, text="Evaluation")
DB_Evaluation = tkinter.StringVar()
DB_Evaluation.set("")
Evaluation_entry = tkinter.Entry(marks_frame, textvariable=DB_Evaluation)
Evaluation_label.grid(row=3, column=1)
Evaluation_entry.grid(row=4, column=1)

Participation_label = tkinter.Label(marks_frame, text="Participation")
DB_Participation = tkinter.StringVar()
DB_Participation.set("")
Participation_entry = tkinter.Entry(marks_frame, textvariable=DB_Participation)
Participation_label.grid(row=3, column=2)
Participation_entry.grid(row=4, column=2)

# Tasks_label = tkinter.Label(marks_frame, text="Tasks")
# DB_Tasks = tkinter.StringVar()
# DB_Tasks.set("")
# Tasks_entry = tkinter.Entry(marks_frame, textvariable=DB_Tasks)
# Tasks_label.grid(row=3, column=3)
# Tasks_entry.grid(row=4, column=3)
#
# Final_label = tkinter.Label(marks_frame, text="Final")
# DB_Final = tkinter.StringVar()
# DB_Final.set("")
# Final_entry = tkinter.Entry(marks_frame, textvariable=DB_Final)
# Final_label.grid(row=2, column=4)
# Final_entry.grid(row=3, column=4)

# NotesEn_label = tkinter.Label(marks_frame, text="NotesEn")
# DB_NotesEn = tkinter.StringVar()
# DB_NotesEn.set("NotesEn DB Value")
# NotesEn_entry = tkinter.Entry(marks_frame, textvariable=DB_NotesEn)
# NotesEn_label.grid(row=1, column=4)
# NotesEn_entry.grid(row=2, column=4)

# NotesAr_label = tkinter.Label(marks_frame, text="NotesAr")
# DB_NotesAr = tkinter.StringVar()
# DB_NotesAr.set("NotesAr DB Value")
# NotesAr_entry = tkinter.Entry(marks_frame, textvariable=DB_NotesAr)
# NotesAr_label.grid(row=3, column=4)
# NotesAr_entry.grid(row=4, column=4)

for widget in marks_frame.winfo_children():
    widget.grid_configure(padx=10, pady=5)

# Accept terms
Review_frame = tkinter.LabelFrame(tab1, text="Confirm your inputs")
Review_frame.grid(row=2, column=0, sticky="news", padx=20, pady=10)

accept_var = tkinter.StringVar(value="Not Accepted")
terms_check = tkinter.Checkbutton(Review_frame,
                                  text="I confirm that I reviewed the inputs and I'm responsible for any changes.",
                                  variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
terms_check.grid(row=2, column=0)

# Button
enter_data_button = tkinter.Button(tab1, text="Enter data", command=enter_data, state="disabled")
enter_data_button.grid(row=3, column=0, sticky="news", padx=20, pady=10)

Clear_All_button = tkinter.Button(tab1, text="Clear All", command=clear_all)
Clear_All_button.grid(row=4, column=0, sticky="news", padx=20, pady=10)

# Students Required Inputs----------------------------------------------------------------------------------------------------------
tab2 = ttk.Frame(tabControl)

tabControl.add(tab2, text='Subjects & Exams')

user_info_frame = tkinter.LabelFrame(tab2, text="Subjects & Exams")
user_info_frame.grid(row=0, column=0, padx=20, pady=20)

StudentId_label2 = tkinter.Label(user_info_frame, text="Student ID")
StudentId_label2.grid(row=0, column=0)
StudentName2_label = tkinter.Label(user_info_frame, text="Student_Name")
StudentName2_label.grid(row=0, column=1)

# Button
St_Data_button2 = tkinter.Button(user_info_frame, text="Get Student Full Data", command=get_Exams_per_student)
St_Data_button2.grid(row=2, column=0, sticky="news", padx=20, pady=10)

StudentId_entry2 = tkinter.Entry(user_info_frame)
DB_StudentName2 = tkinter.StringVar()
DB_StudentName2.set("")
StudentName_value2 = tkinter.Entry(user_info_frame, textvariable=DB_StudentName2, state="normal")
StudentId_entry2.grid(row=1, column=0)
StudentName_value2.grid(row=1, column=1)

GradeId_label2 = tkinter.Label(user_info_frame, text="Grade")
DB_GradeId2 = tkinter.StringVar()
DB_GradeId2.set("")
GradeId_value2 = tkinter.Entry(user_info_frame, textvariable=DB_GradeId2, state="normal")
GradeId_label2.grid(row=0, column=2)
GradeId_value2.grid(row=1, column=2)

Class_label2 = tkinter.Label(user_info_frame, text="Class")
DB_ClassId2 = tkinter.StringVar()
DB_ClassId2.set("")
Class_value2 = tkinter.Entry(user_info_frame, textvariable=DB_ClassId2, state="normal")
Class_label2.grid(row=0, column=3)
Class_value2.grid(row=1, column=3)

LanguageType_label2 = tkinter.Label(user_info_frame, text="LanguageType")
DB_LanguageType2 = tkinter.StringVar()
DB_LanguageType2.set("")
LanguageType_value2 = tkinter.Entry(user_info_frame, textvariable=DB_LanguageType2, state="normal")
LanguageType_label2.grid(row=0, column=4)
LanguageType_value2.grid(row=1, column=4)

# Month_label2 = tkinter.Label(user_info_frame, text="Month")
# Month_Combobox2 = ttk.Combobox(user_info_frame, values=["1", "2", "3", "4", "5", "6"], state="readonly")
# Month_label2.grid(row=2, column=3)
# Month_Combobox2.grid(row=3, column=3)

SubjectName_label2 = tkinter.Label(user_info_frame, text="Subjects")
SubjectName_listbox = tkinter.Listbox(user_info_frame, width=20, height=17)
SubjectName_label2.grid(row=5, column=0)
SubjectName_listbox.grid(row=6, column=0)

SubjectId_label2 = tkinter.Label(user_info_frame, text="Subjects IDs")
SubjectId_listbox = tkinter.Listbox(user_info_frame, width=20, height=17)
SubjectId_label2.grid(row=5, column=1)
SubjectId_listbox.grid(row=6, column=1)

Quizes_label = tkinter.Label(user_info_frame, text="Quizzes Types per grade")
Quizes_listbox = tkinter.Listbox(user_info_frame, width=20, height=17)
Quizes_label.grid(row=5, column=2)
Quizes_listbox.grid(row=6, column=2)

FullMarks_label = tkinter.Label(user_info_frame, text="Full Mark")
FullMarks_listbox = tkinter.Listbox(user_info_frame, width=20, height=17)
FullMarks_label.grid(row=5, column=3)
FullMarks_listbox.grid(row=6, column=3)

RowsNumber_label = tkinter.Label(user_info_frame, text="Number of rows required\n for student for each month")
RowsCount = tkinter.StringVar()
RowsCount.set("")
RowsNumber = tkinter.Entry(user_info_frame, textvariable=RowsCount, state="normal")
RowsNumber_label.grid(row=8, column=0)
RowsNumber.grid(row=9, column=0)

RowsNumber1_1_label = tkinter.Label(user_info_frame, text="Rows for Term: 1, Month: 1")
RowsCount1_1 = tkinter.StringVar()
RowsCount1_1.set("")
RowsNumber1_1 = tkinter.Entry(user_info_frame, textvariable=RowsCount1_1, state="normal")
RowsNumber1_1_label.grid(row=7, column=1)
RowsNumber1_1.grid(row=8, column=1)

RowsNumber1_2_label = tkinter.Label(user_info_frame, text="Rows for Term: 1, Month: 2")
RowsCount1_2 = tkinter.StringVar()
RowsCount1_2.set("")
RowsNumber1_2 = tkinter.Entry(user_info_frame, textvariable=RowsCount1_2, state="normal")
RowsNumber1_2_label.grid(row=7, column=2)
RowsNumber1_2.grid(row=8, column=2)

RowsNumber1_3_label = tkinter.Label(user_info_frame, text="Rows for Term: 1, Month: 3")
RowsCount1_3 = tkinter.StringVar()
RowsCount1_3.set("")
RowsNumber1_3 = tkinter.Entry(user_info_frame, textvariable=RowsCount1_3, state="normal")
RowsNumber1_3_label.grid(row=7, column=3)
RowsNumber1_3.grid(row=8, column=3)

RowsNumber2_1_label = tkinter.Label(user_info_frame, text="Rows for Term: 2, Month: 1")
RowsCount2_1 = tkinter.StringVar()
RowsCount2_1.set("")
RowsNumber2_1 = tkinter.Entry(user_info_frame, textvariable=RowsCount2_1, state="normal")
RowsNumber2_1_label.grid(row=9, column=1)
RowsNumber2_1.grid(row=10, column=1)

RowsNumber2_2_label = tkinter.Label(user_info_frame, text="Rows for Term: 2, Month: 2")
RowsCount2_2 = tkinter.StringVar()
RowsCount2_2.set("")
RowsNumber2_2 = tkinter.Entry(user_info_frame, textvariable=RowsCount2_2, state="normal")
RowsNumber2_2_label.grid(row=9, column=2)
RowsNumber2_2.grid(row=10, column=2)

RowsNumber2_3_label = tkinter.Label(user_info_frame, text="Rows for Term: 2, Month: 3")
RowsCount2_3 = tkinter.StringVar()
RowsCount2_3.set("")
RowsNumber2_3 = tkinter.Entry(user_info_frame, textvariable=RowsCount2_3, state="normal")
RowsNumber2_3_label.grid(row=9, column=3)
RowsNumber2_3.grid(row=10, column=3)
# Button
# get_subject_Id_button2 = tkinter.Button(user_info_frame, text="Get Subject IDs", command=get_subject_Id, state='disabled')
# get_subject_Id_button2.grid(row=0, column=4, sticky="news", padx=20, pady=10)
# SubjectId_label = tkinter.Label(user_info_frame, text="Subject ID")

# DB_SubjectId2 = tkinter.StringVar()
# DB_SubjectId2.set("")
# SubjectId_entry2 = tkinter.Entry(user_info_frame, textvariable=DB_SubjectId2, state="disabled")
# # SubjectId_label.grid(row=0, column=4)
# SubjectId_entry2.grid(row=1, column=4)

# Term_label2 = tkinter.Label(user_info_frame, text="Term")
# Term_combobox2 = ttk.Combobox(user_info_frame, values=["1", "2"], state="readonly")
# Term_label2.grid(row=2, column=4)
# Term_combobox2.grid(row=3, column=4)


# Fill Exams Sheet & Split sheets per subjects tab----------------------------------------------------------------------------------------------------------
tab3 = ttk.Frame(tabControl)
tabControl.add(tab3, text='Fill Exams Sheet & Split sheets per subjects')

TermExtract_label = tkinter.Label(tab3, text="Term")
TermExtract_combobox = ttk.Combobox(tab3, values=["1", "2"], state="readonly")
TermExtract_label.grid(row=0, column=0)
TermExtract_combobox.grid(row=1, column=0)

output_label = tkinter.Label(tab3,
                             text="Output File name (Without extension), file with exist under this path:" + os.getcwd())
output_label.grid(row=3, column=0)
output_entry = tkinter.Entry(tab3)
output_entry.grid(row=4, column=0)

accept_var2 = tkinter.StringVar(value="Not Accepted")
terms_check = tkinter.Checkbutton(tab3,
                                  text="I confirm that I want to create new excel with the generated data.",
                                  variable=accept_var2, onvalue="Accepted", offvalue="Not Accepted")
terms_check.grid(row=7, column=0)

# Button
fill_exams_sheet_button = tkinter.Button(tab3, text="fill Exams sheet", command=fill_exams_sheet, state="normal")
fill_exams_sheet_button.grid(row=8, column=0, sticky="news", padx=20, pady=10)

Create_Subjects_Sheet_label = tkinter.Label(tab3,
                                            text="After creating Exams sheet, Click Start to Create sheet for each subject\n please note that process will complete when progress bar stops loading\n Sheets can be found ")
Create_Subjects_Sheet_label.grid(row=10, column=0)

progress_bar = ttk.Progressbar(tab3, orient='horizontal', length=1150, mode='determinate')
start_button = tkinter.Button(tab3, text='Start', command=Create_sheets_per_subject, state="disabled")
progress_bar.grid(row=11, column=0)
start_button.grid(row=12, column=0)

# Created_Sheets_label = tkinter.Label(tab3, text="Created Sheets")
# Created_Sheets_listbox = tkinter.Listbox(tab3, width=60, height=20)
# Created_Sheets_label.grid(row=7, column=0)
# Created_Sheets_listbox.grid(row=8, column=0)

# Combine sheets after filling marks----------------------------------------------------------------------------------------------------------
tab4 = ttk.Frame(tabControl)
tabControl.add(tab4, text='Combine sheets after filling marks')

Hint_label = tkinter.Label(tab4,
                           text="Important Note, this process should be done after running Split sheets per subjects and fill all marks, then put the sheets back each one in it's place")
Hint_label.grid(row=0, column=0)

TermFolder_label = tkinter.Label(tab4, text="Term Folder")
TermFolder_combobox = ttk.Combobox(tab4, values=["Term_1", "Term_2"], state="readonly")
TermFolder_label.grid(row=1, column=0)
TermFolder_combobox.grid(row=2, column=0)

Create_Final_Marks_sheet_button = tkinter.Button(tab4, text="Combine sheets", command=combine_subjects_sheets,
                                                 state="normal")
Create_Final_Marks_sheet_button.grid(row=3, column=0, sticky="news", padx=20, pady=10)

Notes_label = tkinter.Label(tab4,
                            text="Final combined data can be found in the following sheet" + os.getcwd() + "/Combined_Data_Term_# , please copy the data into Marks sheet in the database file")
Notes_label.grid(row=4, column=0)

window.mainloop()
